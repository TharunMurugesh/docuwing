"""XLSX Parser using openpyxl."""

from __future__ import annotations

import io
from typing import Any

import openpyxl

from docuwing_engine.domain.entities import Document
from docuwing_engine.interfaces.parser import ParserPlugin
from docuwing_engine.ir.types import DocumentIR, Section, TableBlock, TableCell, TableRow
from docuwing_engine.plugins.registry import PluginCategory, PluginManifest


class XlsxParser(ParserPlugin):
    """Parses Excel spreadsheets using openpyxl."""

    MANIFEST = PluginManifest(
        name="xlsx_parser",
        category=PluginCategory.PARSER,
        description="Native parser for XLSX files",
        mime_types=["application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"],
    )

    def initialize(self, config: dict[str, Any] | None = None) -> None:
        pass

    async def parse(self, document: Document, stream: io.BytesIO) -> DocumentIR:
        wb = openpyxl.load_workbook(stream, data_only=True)

        sections = []

        for sheet_name in wb.sheetnames:
            sheet = wb[sheet_name]

            rows = []
            for row in sheet.iter_rows(values_only=True):
                # Skip entirely empty rows
                if all(cell is None for cell in row):
                    continue

                cells = [
                    TableCell(text=str(cell).strip() if cell is not None else "") for cell in row
                ]
                rows.append(TableRow(cells=cells))

            if rows:
                table_block = TableBlock(rows=rows)
                sections.append(Section(title=sheet_name, blocks=[table_block]))

        ir = DocumentIR(
            document_id=document.id,
            pages=len(wb.sheetnames),
            sections=sections,
        )

        return ir
